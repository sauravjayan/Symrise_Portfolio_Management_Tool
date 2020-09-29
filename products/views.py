from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, TemplateView, UpdateView, DeleteView, CreateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from .models import Product
from django.urls import reverse, reverse_lazy
from datetime import datetime
from .forms import ProductForm, QuantityForm, FilterForm, NIFilterForm
from django.http import HttpResponseRedirect, HttpResponse
import re
from openpyxl import Workbook
from django.contrib import messages


# Create your views here.

class ProductTemplateView(LoginRequiredMixin, TemplateView):

    template_name = 'products/products_home.html'

    def post(self, request, *args, **kwargs):
        six_d_code = []
        for pd in Product.objects.all():
            six_d_code.append(str(pd.six_digit_code))

        if request.POST.get('search') in six_d_code:
            return redirect(reverse('products:product_detail', kwargs={'slug':request.POST.get('search')}))
        elif len(Product.objects.filter(name__icontains=request.POST.get('search'))):
            return redirect(reverse('products:uni_search_list', kwargs={'name':request.POST.get('search')}))
        else:
            messages.error(request, 'The item(s) you searched for do(es) not exist.')
            return redirect(reverse('products:home'))
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pd_list = Product.objects.all()
        alert = False
        for pd in pd_list:
            if pd.expiry_alert():
                alert = True
                break
        num = 0
        for pd in pd_list:
            if pd.expiry_alert():
                num += 1
            
        context['pd_list'] = Product.objects.all()
        context['alert'] = alert
        context['num'] = num
        return context


class ProductDetailView(LoginRequiredMixin, DetailView):
    model = Product
    context_object_name = 'product'
    template_name = 'products/products_detail.html'


class NaturalProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'products/products_list.html'
    context_object_name = 'pd_list'

    def post(self, request, *args, **kwargs):
        if request.POST.get('submit') == 'Search':
            return redirect(reverse('products:search_list', kwargs={'name':request.POST.get('search')}))
        
        if request.POST.get('submit') == 'Export':
            return redirect(reverse('products:export', kwargs={'qset':'natural'}))


class NaturalIdenticalProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'products/ni_products_list.html'
    context_object_name = 'pd_list'

    def post(self, request, *args, **kwargs):
        if request.POST.get('submit') == 'Search':
            return redirect(reverse('products:ni_search_list', kwargs={'name':request.POST.get('search')}))
        
        if request.POST.get('submit') == 'Export':
            return redirect(reverse('products:export', kwargs={'qset':'ni'}))


class ProductCreateView(LoginRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'products/product_create_form.html'


class UniversalSearchListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'products/uni_search_list.html'
    context_object_name = 'pd_list'

    def get_queryset(self):
        search_term = self.kwargs.get('name')
        if search_term.isdigit():
            queryset = Product.objects.filter(six_digit_code__iexact=self.kwargs.get('name'))
        else:
            queryset = Product.objects.filter(name__icontains=self.kwargs.get('name'))
        
        if len(queryset) == 0:
            messages.error(self.request, 'The item(s) you searched for do(es) not exist.')

        return queryset
    
    def post(self, request, *args, **kwargs):
        if request.POST.get('submit') == 'Search':
            return redirect(reverse('products:uni_search_list', kwargs={'name':request.POST.get('search')}))


class ProductSearchListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'products/products_list.html'
    context_object_name = 'pd_list'

    def get_queryset(self):
        search_term = self.kwargs.get('name')
        if search_term.isdigit():
            queryset = Product.objects.filter(six_digit_code__iexact=self.kwargs.get('name')).filter(legal_name__icontains='natural')
        else:
            queryset = Product.objects.filter(name__icontains=self.kwargs.get('name')).filter(legal_name__icontains='natural')
        
        if len(queryset) == 0:
            messages.error(self.request, 'The item(s) you searched for do(es) not exist.')
        
        return queryset
    
    def post(self, request, *args, **kwargs):
        if request.POST.get('submit') == 'Search':
            return redirect(reverse('products:search_list', kwargs={'name':request.POST.get('search')}))
        
        if request.POST.get('submit') == 'Export':
            return redirect(reverse('products:export', kwargs={'qset':'natural'}))

class NIProductSearchListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'products/ni_products_list.html'
    context_object_name = 'pd_list'

    def get_queryset(self):
        search_term = self.kwargs.get('name')
        if search_term.isdigit():
            queryset = Product.objects.filter(six_digit_code__iexact=self.kwargs.get('name')).exclude(legal_name__icontains='natural')
        else:
            queryset = Product.objects.filter(name__icontains=self.kwargs.get('name')).exclude(legal_name__icontains='natural')
        
        if len(queryset) == 0:
            messages.error(self.request, 'The item(s) you searched for do(es) not exist.')
        
        return queryset
    
    def post(self, request, *args, **kwargs):
        if request.POST.get('submit') == 'Search':
            return redirect(reverse('products:ni_search_list', kwargs={'name':request.POST.get('search')}))
        
        if request.POST.get('submit') == 'Export':
            return redirect(reverse('products:export', kwargs={'qset':'ni'}))


class ProductDeleteView(LoginRequiredMixin, DeleteView):
    model = Product
    success_url = reverse_lazy('products:home')


class ProductUpdateView(LoginRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'products/product_update_form.html'

class QuantityUpdateView(LoginRequiredMixin, UpdateView):
    model = Product
    form_class = QuantityForm
    template_name = 'products/quantity_update_form.html'


@login_required
def filter(request):

    if request.method == 'POST':
        form = FilterForm(request.POST)

        if form.is_valid():
            f = form.cleaned_data.get('flavour_key')
            a = form.cleaned_data.get('alcohol_content')
            p = form.cleaned_data.get('production_site')
            s = form.cleaned_data.get('solubility')

            return HttpResponseRedirect(reverse('products:filter_list', kwargs={'flavour_key':f, 'alcohol_content':a, 'production_site':p, 'solubility':s }))

    else:
        context = {}
        form = FilterForm()
        context['form'] = form
        return render(request, 'products/filter.html', context)


class FilterListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'products/filter_list.html'
    context_object_name = 'filter_pd'

    def get_queryset(self):
        flv = self.kwargs.get('flavour_key')
        alc = self.kwargs.get('alcohol_content')
        prod = self.kwargs.get('production_site')
        sol = self.kwargs.get('solubility')

        flv_ls = re.findall(r'[A-Za-z]+\s[A-Za-z]+|[A-Za-z]+', flv)
        sol_ls = re.findall(r'[A-Za-z]+\s[a-z]+', sol)
        prod_ls = re.findall(r'[A-Za-z]+\s[A-Za-z]+|[A-Za-z]+',prod)

        if 'Select all' in flv_ls:
            flv_qset = Product.objects.all()
        else:
            flv_qset = Product.objects.filter(flavour_key__icontains=flv_ls[0])
            for i in range(1,len(flv_ls)):
                flv_qset = flv_qset | Product.objects.filter(flavour_key__icontains=flv_ls[i])

        if 'Select all' in prod_ls:
            prod_qset = Product.objects.all()
        else:
            prod_qset = Product.objects.filter(production_site__in=prod_ls)
        
        if 'Select all' in sol_ls:
            sol_qset = Product.objects.all()
        else:
            sol_qset = Product.objects.filter(solubility__in=sol_ls)


        if alc == 'less':
            alc_qset = Product.objects.filter(alcohol_content__lte=0.1)
        elif alc == 'greater':
            alc_qset = Product.objects.filter(alcohol_content__gte=0.1)
        else:
            alc_qset = Product.objects.all()

        queryset = Product.objects.all().intersection(flv_qset, prod_qset, sol_qset, alc_qset)
        
        return queryset    


@login_required
def ni_filter(request):

    if request.method == 'POST':
        form = NIFilterForm(request.POST)

        if form.is_valid():
            f = form.cleaned_data.get('flavour_key')
            a = form.cleaned_data.get('alcohol_content')
            p = form.cleaned_data.get('production_site')
            s = form.cleaned_data.get('solubility')

            return HttpResponseRedirect(reverse('products:ni_filter_list', kwargs={'flavour_key':f, 'alcohol_content':a, 'production_site':p, 'solubility':s }))

    else:
        context = {}
        form = NIFilterForm()
        context['form'] = form
        return render(request, 'products/ni_filter.html', context)


class NIFilterListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'products/ni_filter_list.html'
    context_object_name = 'filter_pd'

    def get_queryset(self):
        flv = self.kwargs.get('flavour_key')
        alc = self.kwargs.get('alcohol_content')
        prod = self.kwargs.get('production_site')
        sol = self.kwargs.get('solubility')

        flv_ls = re.findall(r'[A-Za-z]+\s[A-Za-z]+|[A-Za-z]+', flv)
        sol_ls = re.findall(r'[A-Za-z]+\s[a-z]+', sol)
        prod_ls = re.findall(r'[A-Za-z]+\s[A-Za-z]+|[A-Za-z]+',prod)

        if 'Select all' in flv_ls:
            flv_qset = Product.objects.all()
        else:
            flv_qset = Product.objects.filter(flavour_key__icontains=flv_ls[0])
            for i in range(1,len(flv_ls)):
                flv_qset = flv_qset | Product.objects.filter(flavour_key__icontains=flv_ls[i])

        if 'Select all' in prod_ls:
            prod_qset = Product.objects.all()
        else:
            prod_qset = Product.objects.filter(production_site__in=prod_ls)
        
        if 'Select all' in sol_ls:
            sol_qset = Product.objects.all()
        else:
            sol_qset = Product.objects.filter(solubility__in=sol_ls)


        if alc == 'less':
            alc_qset = Product.objects.filter(alcohol_content__lte=0.1)
        elif alc == 'greater':
            alc_qset = Product.objects.filter(alcohol_content__gte=0.1)
        else:
            alc_qset = Product.objects.all()

        queryset = Product.objects.all().intersection(flv_qset, prod_qset, sol_qset, alc_qset)
        
        return queryset    


@login_required
def export_to_xlsx(request, **kwargs):
    
    if kwargs['qset'] == 'natural':
        product_qset = Product.objects.filter(legal_name__icontains='natural')
    elif kwargs['qset'] == 'ni':
        product_qset = Product.objects.exclude(legal_name__icontains='natural')
    else:
        product_qset = Product.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-products.xlsx'.format(
        date=datetime.now().strftime('%d-%m-%Y'),
    )

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Products'

    columns = [
        'Product Code',
        'Product Name',
        'WLC',
        'Flavour Key',
        'Expiry Date',
        'Order Date',
        'Arrival Date',
        'Lab Location',
        'Quantity',
        'Physical Form',
        'Solubility',
        'Shelf Life',
        'Storage Condition',
        'Production Site',
        'Legal Name',
        'Alcohol Content',
        'Halal Status',
        'Sales Status',
        'Portfolio Manager',

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for pd in product_qset:
        row_num += 1
        wlc = 'Yes'
        if not pd.wlc:
            wlc = 'No'
        
        # Define the data for each cell in the row 
        row = [
           pd.six_digit_code,
           pd.name,
           wlc,
           pd.flavour_key,
           pd.expiry_date,
           pd.order_date,
           pd.arrival_date,
           pd.lab_location,
           pd.quantity,
           pd.physical_form,
           pd.solubility,
           pd.shelf_life,
           pd.storage_condition,
           pd.production_site,
           pd.legal_name,
           pd.alcohol_content,
           pd.halal_status,
           pd.sales_status,
           pd.portfolio_manager,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response