from products.models import Product 
from openpyxl import load_workbook
from datetime import datetime



wb = load_workbook('/Users/sauravjayan/Desktop/SYMRISE_DUBAI_Portfolio_BEV_new.xlsx')
sheet = wb.worksheets[1]

for i in range(4,189):
    pd_list = []
    for j in range(1,19):
        pd_list.append(sheet.cell(row=i, column=j).value)
    
    if pd_list[0] == 'Not written':
        pd_list[0] = None

    if isinstance(pd_list[1], str):
        pd_list[1] = datetime.strptime(pd_list[1], '%d-%m-%Y')
    
    if isinstance(pd_list[2], str):
        pd_list[2] = datetime.strptime(pd_list[2], '%d/%m/%Y')
    
    
    
    if pd_list[4] == 'X':
        pd_list[4] = True
    else:
        pd_list[4] = False

    pd = Product.objects.create(wlc=pd_list[4], six_digit_code=pd_list[5], name=pd_list[6], flavour_key=pd_list[7],
                                        expiry_date=pd_list[0], order_date=pd_list[1], arrival_date=pd_list[2], lab_location=pd_list[3],
                                        physical_form=pd_list[8], solubility=pd_list[9], shelf_life=pd_list[10], storage_condition=pd_list[11], production_site=pd_list[12],
                                        legal_name=pd_list[13], alcohol_content=pd_list[14], halal_status=pd_list[15],
                                        sales_status=pd_list[16], portfolio_manager=pd_list[17])
    pd.save()


